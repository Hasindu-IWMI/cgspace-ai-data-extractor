import pandas as pd
import openpyxl
import logging
import os
from api_client import APIClient
from pdf_processor import PDFProcessor
from config import Config
from utils import interruptable_sleep, chunk_text_safe  # If needed
import tempfile

class ExcelWriter:
    def __init__(self, config):
        self.config = config

    def normalize_col_name(self, col):
        return col.lower().replace(' ', '_').replace('/', '_').replace('-', '_')

    def extract_chunks_and_save(self, query, start_page, end_page, driver, selected_base_fields, min_year, max_year, selected_affiliations, selected_regions, selected_countries, progress_queue):
        chunks_excel = "cgspace_chunks_data.xlsx"
        total_items = APIClient(self.config).get_total_count(query, min_year, max_year, selected_affiliations, selected_regions, selected_countries)
        expected_samples = (end_page - start_page + 1) * 10
        total_samples = min(total_items - (start_page - 1) * 10, expected_samples)
        num_pages = end_page - start_page + 1
        progress_queue.put(f"Total documents for chunking: {total_samples}")
        progress_queue.put(f"Total pages for chunking: {num_pages}") # New: Send total pages for UI
        if total_samples <= 0:
            progress_queue.put("No documents found for chunking.")
            return
        chunk_rows = []
        for page in range(start_page, end_page + 1):
            if not self.config.running_event.is_set():
                progress_queue.put("Stopped chunk extraction.")
                break
            progress_queue.put(f"Processing page {page} of {end_page}") # New: Page progress message (absolute page, but UI will adjust)
            items, _ = APIClient(self.config).search_items(query, page, min_year, max_year, selected_affiliations, selected_regions, selected_countries)
            for i, item in enumerate(items):
                current_item = i + 1
                progress_queue.put(f"Chunking document {current_item} on page {page}")
                item_data, bitstreams, metadata = APIClient(self.config).get_item_details_safe(item, driver)
                if not item_data or not metadata:
                    progress_queue.put(f"Skipped document on page {page}")
                    continue
                item_id = item_data.get("uuid")
                base_metadata = [metadata.get(self.config.FIELD_MAPPING.get(field, field), "Unknown") for field in selected_base_fields]
                pdf_text = ""
                for bitstream in bitstreams[:1]: # Assume first PDF
                    pdf_url = bitstream.get("_links", {}).get("content", {}).get("href")
                    if pdf_url:
                        response = APIClient(self.config).make_api_request_safe(pdf_url, stream=True)
                        total_size = int(response.headers.get('content-length', 0))
                        downloaded_size = 0
                        tmp_path = tempfile.mktemp(suffix='.pdf')
                        with open(tmp_path, 'wb') as f:
                            for data in response.iter_content(8192):
                                f.write(data)
                                downloaded_size += len(data)
                                if total_size > 0:
                                    progress_percent = int((downloaded_size / total_size) * 100)
                                    progress_queue.put(f"Downloading PDF for item {item_id}: {progress_percent}%") # New: Download progress
                                else:
                                    progress_queue.put(f"Downloading PDF for item {item_id}: {downloaded_size} bytes")
                        progress_queue.put(f"Download completed for item {item_id}") # New: Completion message
                        pdf_text = PDFProcessor(self.config).extract_pdf_text_safe(tmp_path, progress_queue=progress_queue, item_id=item_id)
                        os.unlink(tmp_path)
                        break
                if not pdf_text:
                    progress_queue.put(f"No PDF found for {item_id}")
                    logging.info(f"No PDF found for item {item_id}")
                    continue
                chunks = chunk_text_safe(pdf_text, progress_queue=progress_queue, item_id=item_id)
                for chunk_num, chunk in enumerate(chunks, 1):
                    row = base_metadata + [item_id, chunk_num, chunk] # Repeat metadata per chunk row
                    chunk_rows.append(row)
                progress_queue.put(f"Chunked {len(chunks)} parts for {item_id}")
        if chunk_rows:
            logging.info(f"Saving {len(chunk_rows)} chunk rows to {chunks_excel}")
            try:
                df = pd.DataFrame(chunk_rows, columns=[self.normalize_col_name(f) for f in selected_base_fields] + ['item_id', 'chunk_number', 'chunk_text'])
                with pd.ExcelWriter(chunks_excel, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Chunks', index=False)
                progress_queue.put(f"Chunks saved to {chunks_excel}")
                logging.info("Chunks Excel saved successfully")
            except Exception as e:
                logging.error(f"Error saving chunks Excel: {e}")
                progress_queue.put(f"Error saving Excel: {str(e)}")
        else:
            progress_queue.put("No chunks extracted.")
            logging.info("No chunks extracted, no Excel saved")

    def save_to_excel(self, data_list, features, progress_queue=None, extract_ai=False, selected_base_fields=None):
        if not data_list:
            logging.info("No new data to save to Excel")
            return
       
        def is_valid_row(row, min_required_non_null):
            non_null_count = sum(1 for val in row if val is not None and val != "" and pd.notna(val))
            return non_null_count >= min_required_non_null
        try:
            # Define base columns and semantic columns
            base_columns = selected_base_fields or ["title", "date", "author_organization", "geography_focus", "type", "source", "url"]
            normalized_base_columns = [self.normalize_col_name(col) for col in base_columns]
            semantic_columns = [self.normalize_col_name(feature[0]) for feature in features] if extract_ai else []
            columns = normalized_base_columns + semantic_columns
            columns = list(dict.fromkeys(columns)) # Remove duplicates, preserve order
            # Validate and prepare new data
            new_data = []
            num_cols = len(columns)
            min_required_non_null = len(normalized_base_columns) # Require at least base columns to be non-null
            for row_idx, row in enumerate(data_list):
                row = row[:num_cols] + [None] * (num_cols - len(row)) if len(row) < num_cols else row[:num_cols]
                non_null_count = sum(1 for v in row[:len(normalized_base_columns)] if v is not None and v != "")
                if non_null_count < min_required_non_null:
                    logging.warning(f"Skipping malformed new row {row_idx + 1}: {row[:10]}... (only {non_null_count}/{min_required_non_null} non-null base values)")
                    continue
                new_row = {}
                for idx, col in enumerate(columns):
                    new_row[col] = row[idx]
                new_data.append(new_row)
                logging.debug(f"Validated new row {row_idx + 1}: {list(new_row.values())[:10]}...")
            if not new_data:
                logging.warning("No valid new rows to save after validation")
                return
            # Load existing workbook if it exists, else create new
            if os.path.exists(self.config.EXCEL_FILE):
                wb = openpyxl.load_workbook(self.config.EXCEL_FILE)
                ws = wb.active
                existing_columns = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
                if existing_columns != columns:
                    logging.warning("Column mismatch detected. Overwriting with new headers.")
                    ws.delete_rows(1, ws.max_row) # Clear existing data if headers differ
                    ws.append(columns)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(columns) # Write header only if new file
            # Append new rows with progress
            appended_count = 0
            total_rows = len(new_data)
            if progress_queue:
                progress_queue.put("Composing Excel: 0%")
            for row_idx, new_row_dict in enumerate(new_data):
                row_values = [new_row_dict.get(col) for col in columns]
                ws.append(row_values)
                appended_count += 1
                if progress_queue and total_rows > 0:
                    progress_percent = int((appended_count / total_rows) * 100)
                    progress_queue.put(f"Composing Excel: {progress_percent}%")
            # Remove unnecessary columns (if any appeared)
            placeholders = {'or', 'even', 'null', 'always', 'iso'}
            unnecessary_columns = [
                idx + 1 for idx, col in enumerate(columns)
                if (len(str(col)) <= 3 or str(col).isdigit() or str(col) in placeholders)
                and col not in normalized_base_columns # Only check against base columns
            ]
            if unnecessary_columns:
                for col_idx in sorted(unnecessary_columns, reverse=True):
                    ws.delete_cols(col_idx)
                logging.info(f"Removed {len(unnecessary_columns)} unnecessary columns")
            # Save
            wb.save(self.config.EXCEL_FILE)
            logging.info(f"Appended {appended_count} new rows to {self.config.EXCEL_FILE}. Total rows: {ws.max_row}")
            if progress_queue:
                progress_queue.put("Excel composition completed")
        except Exception as e:
            logging.error(f"Failed to save to {self.config.EXCEL_FILE}: {e}")
            raise